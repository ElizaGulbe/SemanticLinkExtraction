import pandas as pd 
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from tqdm import tqdm
import re
import numpy as np


def find_largest_probability(row):
    probabilities = {
        "hypernym": row["hypernym"],
       # "hyponym": row["hyponym"],
        "synonym": row["synonymy"],
        "none": row["none"]
    }
    # Return the key (label) with the maximum value
    return max(probabilities, key=probabilities.get)


def apply_bold_formatting(text):
    """Apply bold formatting to parts of the text in a cell."""
    parts = re.split(r'(\*\*.*?\*\*)', text)
    rich_text = CellRichText()

    for part in parts:
        if part.startswith('*') and part.endswith('*'):
            part_text = part[2:-2]
            rich_text.append(TextBlock(InlineFont(b=True), part_text))
        else:
            rich_text.append(part)
    
    return rich_text

def convert_hyponym_to_hypernym(df): 
    
    # Step 2: Identify the rows where 'hyponym' is equal to the maximum value
    rows_with_max_hyponym = df[df['max_label'] == "hyponym"].index

    # Step 3: List of column pairs to swap between 'sense1_*' and 'sense2_*'
    columns_to_swap = [
        'sense1_id', 'sense1_entry_id', 'sense1_heading', 'sense1_gloss',
        'sense1_pos', 'sense1_parent_sense_id', 'sense1_order_no', 'sense1_tag',
        'sense1_hidden', 'sense1_heading_occurance', 'sense1_lexemes_extra',
    ]

    # Step 4: Loop over the sense1_* and sense2_* pairs and swap them
    for col in columns_to_swap:
            sense1_col = col
            sense2_col = col.replace('sense1', 'sense2')
            # Swap sense1 and sense2 columns for rows with maximum hyponym
            temp = df.loc[rows_with_max_hyponym, sense1_col].copy()
            df.loc[rows_with_max_hyponym, sense1_col] = df.loc[rows_with_max_hyponym, sense2_col]
            df.loc[rows_with_max_hyponym, sense2_col] = temp
    temp_hypernym = df.loc[rows_with_max_hyponym, 'hypernym'].copy()
    df.loc[rows_with_max_hyponym, 'hypernym'] = df.loc[rows_with_max_hyponym, 'hyponym']
    df.loc[rows_with_max_hyponym, 'hyponym'] = temp_hypernym
    df.loc[rows_with_max_hyponym,"max_label"] = "hypernym"
    return df

def get_all_relationships(df):
    dict_relationships = {
        "hypernym": [],
        "synonym": [],
    }

    # Filter out rows where the max_label is "none"
    df = df[df["max_label"] != "none"]

    # Iterate over the dataframe and populate the dictionary
    for idx, row in df.iterrows():

    # Check if the resulting DataFrame is empty
        if row["sense2_heading"] == "reiboņa":
            j = "hello"
        if row["max_label"] == "hypernym": 
            highest_probabilities = df[(df['sense1_entry_id'] == row["sense1_entry_id"]) & 
                                       (df['sense2_entry_id'] == row["sense2_entry_id"]) & 
                                       (df["max_label"] == "hypernym")].sort_values("hypernym",ascending=False).reset_index(drop=True)
            if highest_probabilities.shape[0] >=2 : 
                if highest_probabilities.loc[0,"hypernym"] == highest_probabilities.loc[1,"hypernym"]: 
                    equal_probabilities_highest = highest_probabilities[highest_probabilities["hypernym"] == highest_probabilities.loc[0,"hypernym"]]
                    sense_ids_equal = equal_probabilities_highest["sense2_id"].tolist()
                    if row["sense2_parent_sense_id"] in sense_ids_equal:
                        continue
                highest_sense_ids = [highest_probabilities.loc[0,"sense1_id"], highest_probabilities.loc[0,"sense2_id"]]
                if row["sense1_id"] not in highest_sense_ids or row["sense2_id"] not in highest_sense_ids:
                    continue        
            dict_relationships[row["max_label"]].append([row["sense1_id"], row["sense2_id"],row["hypernym"]])
        elif row["max_label"] == "synonym" :
            highest_probabilities = df[
            ((df['sense1_entry_id'] == row["sense1_entry_id"]) & 
            (df['sense2_entry_id'] == row["sense2_entry_id"]) & 
            (df["max_label"] == "synonym")) |
            ((df['sense1_entry_id'] == row["sense2_entry_id"]) & 
            (df['sense2_entry_id'] == row["sense1_entry_id"]) & 
            (df["max_label"] == "synonym"))
        ].sort_values("synonymy",ascending=False).reset_index(drop=True)
            if highest_probabilities.shape[0] >=2 : 
                if highest_probabilities.loc[0,"synonymy"] == highest_probabilities.loc[1,"synonymy"]: 
                    equal_probabilities_highest = highest_probabilities[highest_probabilities["synonymy"] == highest_probabilities.loc[0,"synonymy"]]
                    sense_ids_equal = set(equal_probabilities_highest["sense2_id"].tolist() + equal_probabilities_highest["sense1_id"].tolist())
                    if row["sense2_parent_sense_id"] in sense_ids_equal or row["sense1_parent_sense_id"] in sense_ids_equal:
                        continue
                highest_sense_ids = [highest_probabilities.loc[0,"sense1_id"], highest_probabilities.loc[0,"sense2_id"]]
                if row["sense1_id"] not in highest_sense_ids or row["sense2_id"] not in highest_sense_ids:
                    continue
           


            dict_relationships[row["max_label"]].append([row["sense1_id"], row["sense2_id"],row["synonymy"]])
            
    # Remove duplicates from each relationship list
    for key in dict_relationships:
        # Convert lists to tuples, remove duplicates by using set, then back to list
        dict_relationships[key] = list(map(list, set(map(tuple, dict_relationships[key]))))

    return dict_relationships



wb = Workbook()
hypernym_sheet = wb.active
hypernym_sheet.title = "Hypernym candidates"
hypernym_sheet.append(["Sense1ID","Word", "Gloss","Sense2ID", "Hypernym candidate word", "Hypernym candidate gloss", "Probability", "Yes/No"])
synonym_sheet = wb.create_sheet(title="Synonym candidates")
synonym_sheet.append(["Word", "Gloss", "Synonym candidate word", "Synonym candidate gloss", "Probability", "Yes/No"])
df = pd.read_excel("Generate candidates/Research paper/current/synonym_research_paper_2.xlsx")
noun_dict = pd.read_csv("/Users/eliza/Desktop/ai.lab programming/Production/Data preprocessing/All nouns/only_nouns_with_senses.csv")

# drop any duplicates 

df = df.drop_duplicates(['sense1_id', 'sense2_id']).reset_index(drop=True)
#Unnamed: 0	sense1_id	sense1_entry_id	sense1_heading	sense1_gloss	
# sense1_pos	sense1_parent_sense_id	sense1_order_no	sense1_tag	
# sense1_hidden	sense1_heading_occurance	sense1_lexemes_extra	
# sense2_id	sense2_entry_id	sense2_heading	sense2_gloss	sense2_pos
# sense2_parent_sense_id	sense2_order_no	sense2_tag	sense2_hidden	
# sense2_heading_occurance	sense2_lexemes_extra	hypernym	hyponym	none	synonymy

df['max_label'] = df.apply(find_largest_probability, axis=1)

#df = convert_hyponym_to_hypernym(df)

unique_relationships_dict = get_all_relationships(df)



dict_numbering = { "a": "1.", "b": "2.", "c":"3.","d":"4.", "e":"5.", "f":"6.", "g":"7.","h":"8.","i":"9.","j":"10.","k":"11.","l":"12.","m":"13.","n":"14.","o":"15.","p":"16.","q":"17.","r":"18.","s":"19.","t":"20.","u":"21.","v":"22.","w":"23.","x":"24.","y":"25.","z":"26."}
trans_table = str.maketrans(dict_numbering)
# sense_id,entry_id,entry_heading,parent_sense_id,
# gloss,order_no,sense_tag,synset_id,hidden
for key, value in unique_relationships_dict.items():
        for list_hypernym in value: 
            sense1_id = list_hypernym[0]
            sense2_id = list_hypernym[1]
            probability = list_hypernym[2]

            try:
                entry1_id = noun_dict.loc[noun_dict["sense_id"] == sense1_id, "entry_id"].iloc[0]
                entry2_id_hyp_candidate = noun_dict.loc[noun_dict["sense_id"] == sense2_id, "entry_id"].iloc[0]
            except:
                continue

            # get all cases where hypernym has been the highest probability and sort it accordingly 
            sense1_gloss = noun_dict[noun_dict["entry_id"] == entry1_id].sort_values(by="sense_tag").reset_index(drop=True)
            sense2_gloss = noun_dict[noun_dict["entry_id"] == entry2_id_hyp_candidate].sort_values(by="sense_tag").reset_index(drop=True)
            sense1_word = sense1_gloss.loc[0,"entry_heading"]
            sense2_word = sense2_gloss.loc[0,"entry_heading"]
            sense1_gloss_list = []
            sense2_gloss_list = []
            for idx, gloss_sense1_row in sense1_gloss.iterrows():
                sense1_tag = gloss_sense1_row["sense_tag"]
                if isinstance(sense1_tag,float): 
                    numbering = "X."
                else:
                    numbering = sense1_tag.replace(sense1_tag[0],dict_numbering[sense1_tag[0]])
                if gloss_sense1_row["sense_id"] == sense1_id:
                    sense1_gloss_list.append(f"**{numbering} {gloss_sense1_row["gloss"]}**")
                else:
                    sense1_gloss_list.append(f"{numbering} {gloss_sense1_row["gloss"]}")

            for idx, gloss_sense2_row in sense2_gloss.iterrows():
                sense2_tag = gloss_sense2_row["sense_tag"]
                if isinstance(sense2_tag,float):
                    numbering = "X."
                else: 
                    numbering = sense2_tag.replace(sense2_tag[0],dict_numbering[sense2_tag[0]])
                if gloss_sense2_row["sense_id"] == sense2_id:
                    sense2_gloss_list.append(f"**{numbering} {gloss_sense2_row["gloss"]}**")
                else:
                    sense2_gloss_list.append(f"{numbering} {gloss_sense2_row["gloss"]}")
            if key == "hypernym":
                hypernym_sheet.append([sense1_id,sense1_word,"\n".join(sense1_gloss_list),sense2_id,sense2_word,"\n".join(sense2_gloss_list),probability])
            elif key == "synonym":
                synonym_sheet.append([sense1_id,sense1_word,"\n".join(sense1_gloss_list),sense2_id,sense2_word,"\n".join(sense2_gloss_list),probability])
            

for sheet in [hypernym_sheet,synonym_sheet]: 
    for cell in sheet[sheet.max_row]:
        cell.alignment = Alignment(wrap_text=True)
        

# Apply bold formatting to the cell values with ⁠ ** ⁠
for sheet in [hypernym_sheet,synonym_sheet]: 
    for row in sheet.iter_rows():
        for cell in row:
            text = cell.value
            if isinstance(text, str):
                cell.value = apply_bold_formatting(text)
        


# Add hyperlinks
underline_font = Font(underline="single", color="0000FF")
for sheet in [hypernym_sheet,synonym_sheet]:
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=5):
        cell_2  = row[0]
        cell_2.hyperlink = "https://tezaurs.lv/" + str(cell_2.value)
        cell_2.style = "Hyperlink"
        cell_2.font = underline_font

        cell_5  = row[3]
        cell_5.hyperlink = "https://tezaurs.lv/" + str(cell_5.value)
        cell_5.style = "Hyperlink"
        cell_5.font = underline_font
    
wb.save("Generate candidates/Research paper/results_6_synonym_only_final.xlsx")



















    
    

